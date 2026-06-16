from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from suu.core.plugins import PluginBase


class XlsxExportPlugin(PluginBase):
    """
    Plugin to export scraped election data as a .xlsx file.
    Produces a styled spreadsheet that can be opened directly in Google Sheets.
    Activated via the --xlsx flag (export_xlsx in context).
    """

    # (key, header label)
    # key=None means it's a derived/formula column with no direct data key
    COLUMNS = [
        ("position",           "Position"),
        ("group",              "Group"),
        ("group_type",         "Group Type"),
        ("candidate_name",     "Candidate Name"),
        ("pronouns",           "Pronouns"),
        ("is_winner",          "Winner?"),
        ("initial_tally",      "Initial Tally"),
        ("final_tally",        "Final Tally"),
        ("image_url",          "Image URL"),
        (None,                 "Photo"),          # =IMAGE() formula column
        ("election_statement", "Election Statement"),
        ("group_link",         "Group Link"),
    ]

    COL_WIDTHS = {
        "position":           30,
        "group":              25,
        "group_type":         15,
        "candidate_name":     25,
        "pronouns":           12,
        "is_winner":          10,
        "initial_tally":      14,
        "final_tally":        12,
        "image_url":          20,
        None:                 18,   # Photo column — wide enough for the formula
        "election_statement": 60,
        "group_link":         20,
    }

    # Row height (points) for data rows — tall enough to show thumbnails
    DATA_ROW_HEIGHT = 80

    def run(self, data: any, context: dict) -> None:
        if not context.get("export_xlsx"):
            return

        if context.get("scrape_type") != "election":
            print("XlsxExportPlugin: Skipping (not an election scrape).")
            return

        positions = data.get("positions", [])
        if not positions:
            print("XlsxExportPlugin: No positions found, skipping.")
            return

        # Build flat row dicts
        rows = []
        for pos in positions:
            for cand in pos.get("winners", []):
                rows.append({
                    "position":           pos.get("title", ""),
                    "group":              pos.get("group", ""),
                    "group_type":         pos.get("group_type", ""),
                    "candidate_name":     cand.get("name", ""),
                    "pronouns":           cand.get("pronouns", ""),
                    "is_winner":          cand.get("is_winner", False),
                    "initial_tally":      cand.get("initial_tally", ""),
                    "final_tally":        cand.get("final_tally", ""),
                    "image_url":          cand.get("image_url") or "",
                    "election_statement": cand.get("election_statement", ""),
                    "group_link":         pos.get("group_link", ""),
                })

        # ------------------------------------------------------------------ #
        # Workbook setup
        # ------------------------------------------------------------------ #
        wb = Workbook()
        ws = wb.active
        ws.title = "Election Results"

        header_font      = Font(bold=True, color="FFFFFF")
        header_fill      = PatternFill(fill_type="solid", fgColor="1A73E8")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # ------------------------------------------------------------------ #
        # Header row
        # ------------------------------------------------------------------ #
        for col_idx, (key, label) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment

        # Work out which column index holds image_url so we can reference it
        # in the IMAGE formula.
        image_url_col_idx = next(
            i for i, (k, _) in enumerate(self.COLUMNS, start=1) if k == "image_url"
        )

        # ------------------------------------------------------------------ #
        # Data rows
        # ------------------------------------------------------------------ #
        for row_idx, row in enumerate(rows, start=2):
            ws.row_dimensions[row_idx].height = self.DATA_ROW_HEIGHT

            for col_idx, (key, _) in enumerate(self.COLUMNS, start=1):
                if key is None:
                    # Formula column: =IMAGE(<image_url cell>)
                    # Mode 2 = fit inside the cell while preserving aspect ratio
                    url_cell_addr = ws.cell(row=row_idx, column=image_url_col_idx).coordinate
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=f'=IMAGE({url_cell_addr},2)')
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    value = row.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(key == "election_statement"),
                    )

        # ------------------------------------------------------------------ #
        # Column widths
        # ------------------------------------------------------------------ #
        for col_idx, (key, _) in enumerate(self.COLUMNS, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = self.COL_WIDTHS.get(key, 20)

        # Freeze header row
        ws.freeze_panes = "A2"

        # ------------------------------------------------------------------ #
        # Filename
        # ------------------------------------------------------------------ #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name_part = ""
        if "election_name" in context:
            safe_name = "_".join(
                filter(
                    None,
                    "".join(
                        c if c.isalnum() else "_" for c in context["election_name"]
                    ).split("_"),
                )
            )[:50]
            name_part = f"_{safe_name.lower()}"

        filename = f"scrape_election{name_part}_{timestamp}.xlsx"

        try:
            wb.save(filename)
            print(f"Saved {len(rows)} rows to {filename}")
            print(f"Tip: upload to Google Drive and open with Sheets — the Photo column will render images inline.")
        except Exception as e:
            print(f"XlsxExportPlugin: Error saving file: {e}")